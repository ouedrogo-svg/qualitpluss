"""Récapitulatif des abonnements approuvés par date réelle de validation."""

from __future__ import annotations

from collections import defaultdict
from decimal import Decimal
from io import BytesIO

from django.urls import reverse

from .models import (
    SubscriptionPlan,
    SubscriptionRequest,
    french_month_name,
)

RECAP_COLUMNS = (
    "Date de validation",
    "Nom",
    "Prénom",
    "Catégorie",
    "Mois d'abonnement",
    "Option d'abonnement",
    "Montant",
)


def _plan_option_label(plan: SubscriptionPlan) -> str:
    return plan.display_label


def _content_month_label(req: SubscriptionRequest) -> str:
    """Mois de contenu auquel l’abonnement donne accès."""
    return req.covered_periods_display()


def _format_amount(amount: Decimal) -> str:
    if amount == amount.to_integral_value():
        return f"{int(amount):,}".replace(",", " ")
    return f"{amount:,.2f}".replace(",", " ")


def _approved_requests_queryset(*, category_ids: set[int] | None = None):
    qs = (
        SubscriptionRequest.objects.filter(
            status=SubscriptionRequest.Status.APPROVED,
            decided_at__isnull=False,
        )
        .select_related("user", "category", "plan")
        .order_by("-decided_at")
    )
    if category_ids is not None:
        if not category_ids:
            return qs.none()
        qs = qs.filter(category_id__in=category_ids)
    return qs


def _row_from_request(req: SubscriptionRequest) -> dict:
    sub_date = req.decided_at.date()
    plan = req.plan
    amount = plan.amount
    return {
        "date": sub_date,
        "date_display": sub_date.strftime("%d/%m/%Y"),
        "nom": (req.user.last_name or "").strip() or "—",
        "prenom": (req.user.first_name or "").strip() or "—",
        "categorie": req.category.name,
        "mois_abonnement": _content_month_label(req),
        "option": _plan_option_label(plan),
        "montant": amount,
        "montant_display": _format_amount(amount),
    }


def subscription_recap_rows(
    *, for_month: tuple[int, int] | None = None, category_ids: set[int] | None = None
) -> list[dict]:
    rows = []
    for req in _approved_requests_queryset(category_ids=category_ids):
        row = _row_from_request(req)
        if for_month is not None:
            year, month = for_month
            if row["date"].year != year or row["date"].month != month:
                continue
        rows.append(row)
    rows.sort(key=lambda r: (r["date"], r["nom"].lower(), r["prenom"].lower()))
    return rows


def _month_total(rows: list[dict]) -> Decimal:
    return sum((r["montant"] for r in rows), Decimal("0"))


def build_subscription_recap_tree(
    *, month_export_url, category_ids: set[int] | None = None
) -> list[dict]:
    """
    Récapitulatif par mois calendaire de validation (decided_at).
    ``month_export_url`` : callable ``(year, month) -> str`` pour le lien d’export du mois.
    """
    months_data: dict[tuple[int, int], list[dict]] = defaultdict(list)

    for req in _approved_requests_queryset(category_ids=category_ids):
        row = _row_from_request(req)
        month_key = (row["date"].year, row["date"].month)
        months_data[month_key].append(row)

    tree = []
    for year, month in sorted(months_data.keys(), reverse=True):
        month_rows = sorted(
            months_data[(year, month)],
            key=lambda r: (r["date"], r["nom"].lower(), r["prenom"].lower()),
        )
        tree.append(
            {
                "year": year,
                "month": month,
                "label": f"{french_month_name(month)} {year}",
                "rows": month_rows,
                "n_subscriptions": len(month_rows),
                "total_amount": _month_total(month_rows),
                "total_display": _format_amount(_month_total(month_rows)),
                "export_url": month_export_url(year, month),
            }
        )
    return tree


def subscription_recap_global_export_url() -> str:
    return reverse("admin:courses_subscriptionrequest_export_recap")


def subscription_recap_month_export_url(year: int, month: int) -> str:
    return reverse(
        "admin:courses_subscriptionrequest_export_recap_month",
        args=[f"{year:04d}-{month:02d}"],
    )


def build_admin_subscription_recap_tree() -> list[dict]:
    """Récap admin : liens d’export vers les vues ``admin``."""
    return build_subscription_recap_tree(month_export_url=subscription_recap_month_export_url)


def formateur_subscription_recap_global_export_url() -> str:
    return reverse("courses:formateur_recap_export_all")


def formateur_subscription_recap_month_export_url(year: int, month: int) -> str:
    return reverse(
        "courses:formateur_recap_export_month",
        args=[f"{year:04d}-{month:02d}"],
    )


def build_formateur_subscription_recap_tree(user) -> list[dict]:
    """Récap pour l’espace formateur complet : URLs d’export du site public."""
    from .formateur_permissions import formateur_category_ids

    return build_subscription_recap_tree(
        month_export_url=formateur_subscription_recap_month_export_url,
        category_ids=formateur_category_ids(user),
    )


def formateur_contenu_subscription_recap_global_export_url() -> str:
    return reverse("courses:formateur_contenu_recap_export_all")


def formateur_contenu_subscription_recap_month_export_url(year: int, month: int) -> str:
    return reverse(
        "courses:formateur_contenu_recap_export_month",
        args=[f"{year:04d}-{month:02d}"],
    )


def build_formateur_contenu_subscription_recap_tree(user) -> list[dict]:
    """Récap pour l’espace formateur contenu (sans gestion des demandes)."""
    from .formateur_permissions import formateur_category_ids

    return build_subscription_recap_tree(
        month_export_url=formateur_contenu_subscription_recap_month_export_url,
        category_ids=formateur_category_ids(user, assigned_only=True),
    )


def subscription_recap_spreadsheet_rows(
    *,
    for_month: tuple[int, int] | None = None,
    category_ids: set[int] | None = None,
) -> list[tuple]:
    return [
        (
            row["date_display"],
            row["nom"],
            row["prenom"],
            row["categorie"],
            row["mois_abonnement"],
            row["option"],
            float(row["montant"]),
        )
        for row in subscription_recap_rows(
            for_month=for_month, category_ids=category_ids
        )
    ]


def subscription_recap_filename(
    *, for_month: tuple[int, int] | None = None, ext: str = "xlsx"
) -> str:
    if for_month is not None:
        year, month = for_month
        return f"abonnements_{year:04d}-{month:02d}.{ext}"
    return f"abonnements_recap.{ext}"


def _recap_title(*, for_month: tuple[int, int] | None = None) -> str:
    if for_month is not None:
        year, month = for_month
        return f"Abonnements approuvés — {french_month_name(month)} {year}"
    return "Récapitulatif des abonnements approuvés"


def build_subscription_recap_xlsx(
    *,
    for_month: tuple[int, int] | None = None,
    category_ids: set[int] | None = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    rows = subscription_recap_spreadsheet_rows(
        for_month=for_month, category_ids=category_ids
    )
    title = _recap_title(for_month=for_month)

    wb = Workbook()
    ws = wb.active
    ws.title = "Abonnements"

    ws.append([title])
    ws.append(
        ["Date réelle d’abonnement = date d’approbation par l’administrateur."]
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(RECAP_COLUMNS))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(RECAP_COLUMNS))
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])

    ws.append(list(RECAP_COLUMNS))
    header_row = ws.max_row
    n_cols = len(RECAP_COLUMNS)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E2E8F0")
        cell.alignment = Alignment(horizontal="center")

    total = Decimal("0")
    for row_data in rows:
        ws.append(list(row_data))
        total += Decimal(str(row_data[-1]))

    data_last_row = ws.max_row
    if rows:
        total_row = data_last_row + 2
        ws.cell(row=total_row, column=n_cols - 1, value="Total")
        ws.cell(row=total_row, column=n_cols - 1).font = Font(bold=True)
        ws.cell(row=total_row, column=n_cols - 1).alignment = Alignment(
            horizontal="right"
        )
        total_cell = ws.cell(row=total_row, column=n_cols, value=float(total))
        total_cell.font = Font(bold=True)
        total_cell.number_format = "#,##0"
        total_cell.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 14

    if rows:
        for row in ws.iter_rows(
            min_row=header_row + 1,
            max_row=data_last_row,
            min_col=n_cols,
            max_col=n_cols,
        ):
            amount_cell = row[0]
            if isinstance(amount_cell.value, (int, float)):
                amount_cell.number_format = "#,##0"
                amount_cell.alignment = Alignment(horizontal="right")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
