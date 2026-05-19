"""Classement des candidats pour un examen (meilleure note par personne)."""

from __future__ import annotations

import re
from io import BytesIO

from django.contrib.auth import get_user_model
from django.db.models import Max

from .models import ExamQuizAttempt, MonthlyExam

User = get_user_model()

RESULT_COLUMNS = ("Nom", "Prénom", "Note", "Classement")


def _exam_display_title(exam: MonthlyExam) -> str:
    title = (exam.title or "").strip()
    if title:
        return f"{exam.category} — {exam.period_label} — {title}"
    return f"{exam.category} — {exam.period_label} (examen #{exam.pk})"


def _rank_rows_from_user_scores(
    scores: dict[int, int], users: dict[int, User]
) -> list[dict]:
    rows = []
    for user_id, note in scores.items():
        user = users.get(user_id)
        if not user:
            continue
        rows.append(
            {
                "nom": (user.last_name or "").strip() or "—",
                "prenom": (user.first_name or "").strip() or "—",
                "note": note,
                "classement": 0,
            }
        )
    rows.sort(key=lambda r: (-r["note"], r["nom"].lower(), r["prenom"].lower()))
    rank = 0
    for index, row in enumerate(rows):
        if index == 0 or row["note"] != rows[index - 1]["note"]:
            rank = index + 1
        row["classement"] = rank
    return rows


def ranked_exam_results(exam: MonthlyExam) -> list[dict]:
    """
    Une ligne par candidat dont la première composition a été transmise à l’administrateur.
    Les reprises du quiz ne figurent pas dans ce récapitulatif.
    """
    if not exam.pk:
        return []

    best_by_user = (
        ExamQuizAttempt.objects.filter(exam=exam, sent_to_admin=True)
        .values("user_id")
        .annotate(note=Max("score_points"))
    )
    if not best_by_user:
        return []

    scores = {row["user_id"]: row["note"] for row in best_by_user}
    users = {
        u.pk: u
        for u in User.objects.filter(pk__in=scores.keys()).only(
            "id", "first_name", "last_name", "username"
        )
    }
    return _rank_rows_from_user_scores(scores, users)


def build_admin_exam_recap_tree() -> list[dict]:
    """
    Récapitulatif hiérarchique pour l’en-tête admin : catégorie → mois → examen → classement.
    """
    from collections import defaultdict

    from django.urls import reverse

    exams = list(
        MonthlyExam.objects.select_related("category").order_by(
            "category__name", "-year", "-month", "id"
        )
    )
    if not exams:
        return []

    exam_ids = [e.pk for e in exams]
    scores_by_exam: dict[int, dict[int, int]] = defaultdict(dict)
    user_ids: set[int] = set()
    for row in (
        ExamQuizAttempt.objects.filter(exam_id__in=exam_ids, sent_to_admin=True)
        .values("exam_id", "user_id")
        .annotate(note=Max("score_points"))
    ):
        scores_by_exam[row["exam_id"]][row["user_id"]] = row["note"]
        user_ids.add(row["user_id"])

    users = {
        u.pk: u
        for u in User.objects.filter(pk__in=user_ids).only(
            "id", "first_name", "last_name", "username"
        )
    }

    categories: dict[str, dict] = {}
    for exam in exams:
        cat_name = exam.category.name
        if cat_name not in categories:
            categories[cat_name] = {"name": cat_name, "months": {}}
        month_key = (exam.year, exam.month)
        months = categories[cat_name]["months"]
        if month_key not in months:
            months[month_key] = {
                "label": exam.period_label(),
                "exams": [],
            }
        exam_scores = scores_by_exam.get(exam.pk, {})
        rows = _rank_rows_from_user_scores(exam_scores, users)
        title = (exam.title or "").strip() or f"Examen #{exam.pk}"
        months[month_key]["exams"].append(
            {
                "title": title,
                "display_title": _exam_display_title(exam),
                "n_candidates": len(rows),
                "rows": rows,
                "change_url": reverse(
                    "admin:courses_monthlyexam_change", args=[exam.pk]
                ),
                "export_url": reverse(
                    "admin:courses_monthlyexam_export_resultats", args=[exam.pk]
                ),
            }
        )

    tree = []
    for cat_name in sorted(categories.keys()):
        cat = categories[cat_name]
        month_list = [
            cat["months"][key]
            for key in sorted(cat["months"].keys(), reverse=True)
        ]
        tree.append({"name": cat_name, "months": month_list})
    return tree


def exam_results_spreadsheet_rows(exam: MonthlyExam) -> list[tuple]:
    """Lignes de données (sans en-tête) pour export tabulaire."""
    return [
        (row["nom"], row["prenom"], row["note"], row["classement"])
        for row in ranked_exam_results(exam)
    ]


def exam_results_filename(exam: MonthlyExam, ext: str = "xlsx") -> str:
    label = (exam.title or "").strip() or f"examen_{exam.pk}"
    safe = re.sub(r"[^\w\-]+", "_", label, flags=re.UNICODE).strip("_") or "examen"
    cat = exam.category.slug if exam.category_id else "categorie"
    return f"resultats_{cat}_{exam.year}_{exam.month:02d}_{safe}.{ext}"


def build_exam_results_xlsx(exam: MonthlyExam) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Résultats"

    ws.append([f"Examen : {_exam_display_title(exam)}"])
    ws.append(["Premières compositions transmises à l’administrateur (hors reprises)."])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(RESULT_COLUMNS))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(RESULT_COLUMNS))
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])

    header_row = len(ws["A"])
    ws.append(list(RESULT_COLUMNS))
    for col in range(1, len(RESULT_COLUMNS) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E2E8F0")
        cell.alignment = Alignment(horizontal="center")

    for nom, prenom, note, classement in exam_results_spreadsheet_rows(exam):
        ws.append([nom, prenom, note, classement])

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14

    for row in ws.iter_rows(min_row=header_row + 1, max_col=4):
        row[2].alignment = Alignment(horizontal="right")
        row[3].alignment = Alignment(horizontal="center")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
